#!/usr/bin/env python3
"""Parse WHO GHE 2021 country tables → seeds/ghe_burden_by_slug.json."""
from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import urllib.request
from pathlib import Path

_ROOT = Path(__file__).parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

SEEDS = Path(__file__).parent / "seeds"
MANIFEST = SEEDS / "disease_db_100_manifest.json"
DALY_XLSX = SEEDS / "ghe2021_daly_bycountry_2021.xlsx"
DEATH_XLSX = SEEDS / "ghe2021_deaths_bycountry_2021.xlsx"
OUT = SEEDS / "ghe_burden_by_slug.json"
MAP_OUT = SEEDS / "ghe_slug_map.json"

GHE_DOWNLOADS = {
    DALY_XLSX: (
        "https://cdn.who.int/media/docs/default-source/gho-documents/"
        "global-health-estimates/ghe2021_daly_bycountry_2021.xlsx"
    ),
    DEATH_XLSX: (
        "https://cdn.who.int/media/docs/default-source/gho-documents/"
        "global-health-estimates/ghe2021_deaths_bycountry_2021.xlsx"
    ),
}

# slug → GHE cause code (see WHO GHE 2021 cause list)
GHE_SLUG_OVERRIDES: dict[str, int] = {
    "type-2-diabetes": 800,
    "type-1-diabetes": 800,
    "essential-hypertension": 1120,
    "rheumatoid-arthritis": 1350,
    "major-depressive-disorder": 830,
    "epilepsy": 970,
    "hepatitis-c": 185,
    "inflammatory-bowel-disease": 1244,
    "asthma": 1190,
    "chronic-obstructive-pulmonary-disease": 1180,
    "gastroesophageal-reflux-disease": 1250,
    "atrial-fibrillation": 1160,
    "osteoarthritis": 1360,
    "alzheimers-disease-and-other-dementias": 950,
    "hypothyroidism": 814,
    "low-back-pain": 1380,
    "migraine-disorder": 990,
    "multiple-sclerosis": 980,
    "metabolic-dysfunction-associated-steatohepatitis": 1230,
    "chronic-kidney-disease": 1270,
    "heart-failure": 1150,
    "psoriasis-vulgaris": 1330,
    "parkinson-disease": 960,
    "hivaids": 100,
    "ischemic-stroke": 1140,
    "lung-cancer": 680,
    "colorectal-cancer": 650,
    "breast-cancer": 700,
    "obesity": 814,
    "osteoporosis": 1390,
    "bipolar-disorder": 840,
    "schizophrenia": 850,
    "attention-deficit-hyperactivity-disorder": 910,
    "alcohol-abuse": 860,
    "anxiety": 880,
    "opioid-use-disorder": 870,
    "atopic-eczema": 1330,
    "insomnia": 930,
    "polycystic-ovary-syndrome": 1320,
    "peripheral-arterial-disease": 1160,
    "sepsis": 370,
    "celiac-disease": 1250,
    "gout": 1370,
    "irritable-bowel-syndrome": 1250,
    "systemic-lupus-erythematosus": 814,
    "myalgic-encephalomyelitis-chronic-fatigue-syndrome": 930,
    "ankylosing-spondylitis": 1390,
    "age-related-macular-degeneration": 1060,
    "glaucoma": 1030,
    "diabetic-neuropathy": 1010,
    "pulmonary-arterial-hypertension": 1160,
    "lyme-disease": 370,
    "hepatitis-b-virus-infection": 185,
    "obstructive-sleep-apnea-syndrome": 1200,
    "endometriosis": 1320,
    "benign-prostatic-hyperplasia": 1280,
    "acne": 1330,
    "cystic-fibrosis": 1200,
    "post-acute-covidvaccination-syndrome": 395,
    "prostate-cancer": 740,
    "non-hodgkin-lymphoma": 760,
    "b-cell-chronic-lymphocytic-leukemia": 770,
    "plasma-cell-myeloma": 760,
    "sickle-cell-disease": 812,
    "psoriatic-arthritis": 1390,
    "lupus-nephritis": 1270,
    "sjögrens-syndrome": 814,
    "hemophilia": 813,
    "chronic-venous-insufficiency": 1160,
    "thyroid-cancer": 754,
    "venous-thromboembolism": 1160,
    "myelodysplastic-syndrome": 770,
    "polycystic-kidney-disease": 1270,
    "dermatomyositis": 1390,
    "alopecia-areata": 1330,
    "vitiligo": 1330,
    "tinnitus": 1090,
    "restless-legs-syndrome": 1010,
    "carpal-tunnel-syndrome": 1390,
    "overactive-bladder": 1300,
    "nephrolithiasis": 1290,
    "gastroparesis": 1250,
    "interstitial-cystitis": 1300,
    "myasthenia-gravis": 1010,
    "amyotrophic-lateral-sclerosis": 1010,
    "huntington-disease": 1010,
    "fibromyalgia": 1390,
    "melanoma": 690,
    "exocrine-pancreatic-carcinoma": 670,
    "ovarian-cancer": 730,
    "urinary-bladder-carcinoma": 750,
    "cervical-cancer": 710,
    "acquired-polycythemia-vera": 770,
    "chronic-pancreatitis": 1248,
    "autoimmune-thrombocytopenic-purpura": 814,
    "chronic-idiopathic-urticaria": 1330,
    "alpha-1-antitrypsin-deficiency": 1200,
    "sclerosing-cholangitis": 1230,
    "cancer": 610,
    "measles": 120,
    "sarcopenia": 1390,
}

log = logging.getLogger(__name__)


def _norm(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _ensure_xlsx(path: Path) -> None:
    if path.exists():
        return
    url = GHE_DOWNLOADS.get(path)
    if not url:
        raise FileNotFoundError(path)
    log.info("Downloading %s", path.name)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as r:
        path.write_bytes(r.read())


def _parse_sheet(path: Path) -> dict[int, dict]:
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["All ages"]
    isos = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
    usa_col = next(i for i, iso in enumerate(isos) if iso == "USA")
    country_cols = [i for i in range(7, len(isos)) if isos[i]]

    causes: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=11, values_only=True):
        row = list(row)
        if not row or row[0] != "Persons":
            continue
        code = row[1]
        if code is None:
            continue
        parts: list[str] = []
        for j in range(2, 6):
            v = row[j]
            if v is None:
                continue
            s = str(v).strip()
            if s.endswith("."):
                s = s[:-1].strip()
            if s and len(s) > 1:
                parts.append(s)
        label = parts[-1] if parts else ""
        usa = row[usa_col] if len(row) > usa_col else None
        world = 0.0
        for ci in country_cols:
            v = row[ci]
            if isinstance(v, (int, float)):
                world += float(v)
        causes[int(code)] = {
            "label": label,
            "full": " > ".join(parts),
            "usa_000": float(usa) if isinstance(usa, (int, float)) else None,
            "world_000": world,
        }
    return causes


def _match_slug(slug: str, label: str, query: str, by_label: dict[str, int]) -> int | None:
    if slug in GHE_SLUG_OVERRIDES:
        return GHE_SLUG_OVERRIDES[slug]
    for text in (query, label):
        n = _norm(text)
        if n in by_label:
            return by_label[n]
    q = _norm(query or label)
    for norm_label, code in by_label.items():
        if len(norm_label) >= 5 and (norm_label in q or q in norm_label):
            return code
    return None


def build_seed(extra_slugs: list[str] | None = None) -> dict:
    _ensure_xlsx(DALY_XLSX)
    _ensure_xlsx(DEATH_XLSX)

    daly = _parse_sheet(DALY_XLSX)
    death = _parse_sheet(DEATH_XLSX)
    by_label = {_norm(v["label"]): code for code, v in daly.items() if v.get("label")}

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    slug_rows: dict[str, dict] = {r["slug"]: r for r in manifest}

    for slug in extra_slugs or []:
        slug_rows.setdefault(slug, {"slug": slug, "label": slug, "query": slug})

    by_slug: dict[str, dict] = {}
    slug_map: dict[str, dict] = {}
    unmapped: list[str] = []

    for slug, row in slug_rows.items():
        code = _match_slug(slug, row.get("label", ""), row.get("query", ""), by_label)
        if code is None or code not in daly:
            unmapped.append(slug)
            continue

        d = daly[code]
        de = death.get(code, {})
        usa_dalys = (d.get("usa_000") or 0) * 1000
        global_dalys = (d.get("world_000") or 0) * 1000
        usa_deaths = (de.get("usa_000") or 0) * 1000
        global_deaths = (de.get("world_000") or 0) * 1000

        entry = {
            "ghe_code": code,
            "ghe_cause": d.get("label"),
            "ghe_cause_full": d.get("full"),
            "ghe_year": 2021,
            "us_dalys": round(usa_dalys, 1),
            "global_dalys": round(global_dalys, 1),
            "us_deaths": round(usa_deaths, 1),
            "global_deaths": round(global_deaths, 1),
            "data_tier": "epidemiological",
            "source": "WHO GHE",
        }
        by_slug[slug] = entry
        slug_map[slug] = {
            "ghe_code": code,
            "ghe_cause": d.get("label"),
            "match": "override" if slug in GHE_SLUG_OVERRIDES else "auto",
        }

    if unmapped:
        log.warning("Unmapped slugs (%d): %s", len(unmapped), ", ".join(unmapped))

    return {
        "source": "WHO GHE 2021",
        "year": 2021,
        "daly_file": DALY_XLSX.name,
        "death_file": DEATH_XLSX.name,
        "mapped": len(by_slug),
        "unmapped": unmapped,
        "by_slug": by_slug,
        "_slug_map": slug_map,
    }


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser()
    parser.add_argument("--extra-slug", action="append", default=[])
    args = parser.parse_args(argv)

    payload = build_seed(extra_slugs=args.extra_slug or ["cancer", "measles", "sarcopenia"])
    slug_map = payload.pop("_slug_map")
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    MAP_OUT.write_text(json.dumps(slug_map, indent=2, ensure_ascii=False), encoding="utf-8")
    log.info("Wrote %d entries to %s", payload["mapped"], OUT)
    return 0 if payload["mapped"] else 1


if __name__ == "__main__":
    raise SystemExit(main())